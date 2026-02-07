"""Zero Migration Friction Engine — Import Service.

Handles Excel/CSV parsing, validation, column mapping,
bulk hotel creation, and image downloading.
"""
from __future__ import annotations

import asyncio
import csv
import io
import logging
import os
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

import aiofiles
import httpx

logger = logging.getLogger(__name__)


def _now() -> datetime:
    return datetime.now(timezone.utc)


# ── Excel / CSV Parsing ────────────────────────────────────────

def parse_excel(file_bytes: bytes, filename: str) -> Tuple[List[str], List[List[str]]]:
    """Parse XLSX or CSV file. Returns (headers, rows)."""
    if filename.lower().endswith((".xlsx", ".xls")):
        return _parse_xlsx(file_bytes)
    elif filename.lower().endswith(".csv"):
        return _parse_csv(file_bytes)
    else:
        raise ValueError(f"Desteklenmeyen dosya formatı: {filename}")


def _parse_xlsx(file_bytes: bytes) -> Tuple[List[str], List[List[str]]]:
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel dosyasında aktif sayfa bulunamadı.")

    rows_raw = []
    for row in ws.iter_rows(values_only=True):
        rows_raw.append([str(cell) if cell is not None else "" for cell in row])

    wb.close()

    if len(rows_raw) < 2:
        raise ValueError("Dosyada en az 1 başlık ve 1 veri satırı olmalı.")

    headers = rows_raw[0]
    data_rows = rows_raw[1:]
    # Filter out completely empty rows
    data_rows = [r for r in data_rows if any(cell.strip() for cell in r)]
    return headers, data_rows


def _parse_csv(file_bytes: bytes) -> Tuple[List[str], List[List[str]]]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows_raw = list(reader)

    if len(rows_raw) < 2:
        raise ValueError("Dosyada en az 1 başlık ve 1 veri satırı olmalı.")

    headers = rows_raw[0]
    data_rows = rows_raw[1:]
    data_rows = [r for r in data_rows if any(cell.strip() for cell in r)]
    return headers, data_rows


# ── Column Mapping ─────────────────────────────────────────────

VALID_FIELDS = ["name", "city", "country", "description", "price", "image_url", "address", "phone", "email", "stars", "ignore"]


def map_columns(headers: List[str], rows: List[List[str]], mapping: Dict[str, str]) -> List[Dict[str, Any]]:
    """Apply column mapping to raw rows.

    mapping: { "0": "name", "1": "city", ... } (index-to-field)
    """
    result = []
    for row in rows:
        record: Dict[str, Any] = {}
        for idx_str, field in mapping.items():
            idx = int(idx_str)
            if field == "ignore" or idx >= len(row):
                continue
            value = row[idx].strip() if idx < len(row) else ""
            record[field] = value
        result.append(record)
    return result


# ── Validation ─────────────────────────────────────────────────

def validate_hotels(
    rows: List[Dict[str, Any]],
    existing_names: set[str],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Validate hotel rows.

    Returns (valid_rows, errors).
    errors = [{ row_number, field, message }]
    """
    valid = []
    errors = []
    seen_names: set[str] = set()

    for i, row in enumerate(rows):
        row_num = i + 2  # +2 because row 1 is header, data starts at 2
        row_errors = []

        name = (row.get("name") or "").strip()
        if not name:
            row_errors.append({"row_number": row_num, "field": "name", "message": "Otel adı zorunlu."})
        else:
            name_lower = name.lower()
            if name_lower in existing_names:
                row_errors.append({"row_number": row_num, "field": "name", "message": f"Bu otel zaten mevcut: {name}"})
            elif name_lower in seen_names:
                row_errors.append({"row_number": row_num, "field": "name", "message": f"Dosyada tekrar eden otel: {name}"})
            else:
                seen_names.add(name_lower)

        city = (row.get("city") or "").strip()
        if not city:
            row_errors.append({"row_number": row_num, "field": "city", "message": "Şehir zorunlu."})

        price = row.get("price", "")
        if price:
            try:
                float(str(price).replace(",", "."))
            except (ValueError, TypeError):
                row_errors.append({"row_number": row_num, "field": "price", "message": f"Geçersiz fiyat: {price}"})

        stars = row.get("stars", "")
        if stars:
            try:
                s = int(str(stars))
                if s < 1 or s > 5:
                    row_errors.append({"row_number": row_num, "field": "stars", "message": "Yıldız 1-5 arası olmalı."})
            except (ValueError, TypeError):
                row_errors.append({"row_number": row_num, "field": "stars", "message": f"Geçersiz yıldız: {stars}"})

        if row_errors:
            errors.extend(row_errors)
        else:
            valid.append({**row, "_row_number": row_num})

    return valid, errors


# ── Duplicate Detection ────────────────────────────────────────

async def get_existing_hotel_names(db, org_id: str) -> set[str]:
    """Get set of existing hotel names (lowercased) for the organization."""
    cursor = db.hotels.find({"organization_id": org_id}, {"name": 1})
    names = set()
    async for doc in cursor:
        n = (doc.get("name") or "").strip().lower()
        if n:
            names.add(n)
    return names


# ── Bulk Insert ────────────────────────────────────────────────

async def create_hotels_bulk(
    db,
    org_id: str,
    rows: List[Dict[str, Any]],
    created_by: str,
    job_id: str,
    batch_size: int = 100,
) -> Tuple[int, int, List[Dict[str, Any]]]:
    """Bulk create hotels in batches.

    Returns (success_count, error_count, errors).
    """
    now = _now()
    success = 0
    error_count = 0
    errors = []

    for batch_start in range(0, len(rows), batch_size):
        batch = rows[batch_start:batch_start + batch_size]
        docs = []
        for row in batch:
            row_num = row.pop("_row_number", 0)
            try:
                price_val = None
                if row.get("price"):
                    try:
                        price_val = float(str(row["price"]).replace(",", "."))
                    except (ValueError, TypeError):
                        pass

                stars_val = None
                if row.get("stars"):
                    try:
                        stars_val = int(str(row["stars"]))
                    except (ValueError, TypeError):
                        pass

                doc = {
                    "_id": str(uuid.uuid4()),
                    "organization_id": org_id,
                    "name": row.get("name", "").strip(),
                    "city": row.get("city", "").strip(),
                    "country": row.get("country", "TR").strip() or "TR",
                    "active": True,
                    "created_at": now,
                    "updated_at": now,
                    "created_by": created_by,
                    "updated_by": created_by,
                    "import_job_id": job_id,
                }
                if row.get("description"):
                    doc["description"] = row["description"].strip()
                if price_val is not None:
                    doc["base_price"] = price_val
                if row.get("address"):
                    doc["address"] = row["address"].strip()
                if row.get("phone"):
                    doc["phone"] = row["phone"].strip()
                if row.get("email"):
                    doc["email"] = row["email"].strip()
                if stars_val is not None:
                    doc["stars"] = stars_val
                if row.get("image_url"):
                    doc["image_url"] = row["image_url"].strip()

                docs.append(doc)
            except Exception as e:
                error_count += 1
                errors.append({"row_number": row_num, "field": "general", "message": str(e)})

        if docs:
            try:
                await db.hotels.insert_many(docs, ordered=False)
                success += len(docs)
            except Exception as e:
                # Some may have succeeded in unordered insert
                logger.error("Batch insert error: %s", e)
                error_count += len(docs)
                errors.append({"row_number": 0, "field": "batch", "message": str(e)})

    return success, error_count, errors


# ── Image Downloader ───────────────────────────────────────────

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads", "hotel_images")


async def download_hotel_images(db, org_id: str, job_id: str, max_retries: int = 2) -> int:
    """Download images for hotels that have image_url from a specific import job.

    Returns count of successfully downloaded images.
    """
    os.makedirs(UPLOAD_DIR, exist_ok=True)

    cursor = db.hotels.find({
        "organization_id": org_id,
        "import_job_id": job_id,
        "image_url": {"$exists": True, "$ne": ""},
    })

    downloaded = 0
    async for hotel in cursor:
        url = hotel.get("image_url", "")
        if not url or not url.startswith("http"):
            continue

        hotel_id = hotel["_id"]
        ext = ".jpg"
        if ".png" in url.lower():
            ext = ".png"
        elif ".webp" in url.lower():
            ext = ".webp"

        filename = f"{hotel_id}{ext}"
        filepath = os.path.join(UPLOAD_DIR, filename)

        for attempt in range(max_retries):
            try:
                async with httpx.AsyncClient(timeout=15.0) as client:
                    resp = await client.get(url)
                    if resp.status_code == 200:
                        async with aiofiles.open(filepath, "wb") as f:
                            await f.write(resp.content)
                        await db.hotels.update_one(
                            {"_id": hotel_id},
                            {"$set": {"local_image": f"/uploads/hotel_images/{filename}"}},
                        )
                        downloaded += 1
                        break
            except Exception as e:
                logger.warning("Image download attempt %d failed for %s: %s", attempt + 1, url, e)
                if attempt < max_retries - 1:
                    await asyncio.sleep(1)

    return downloaded


# ── Import Job Lifecycle ───────────────────────────────────────

async def create_import_job(
    db,
    *,
    tenant_id: str,
    organization_id: str,
    entity_type: str = "hotel",
    source: str = "excel",
    total_rows: int = 0,
    filename: str = "",
) -> Dict[str, Any]:
    doc = {
        "_id": str(uuid.uuid4()),
        "tenant_id": tenant_id,
        "organization_id": organization_id,
        "entity_type": entity_type,
        "source": source,
        "status": "uploaded",
        "filename": filename,
        "total_rows": total_rows,
        "success_count": 0,
        "error_count": 0,
        "created_at": _now(),
        "completed_at": None,
    }
    await db.import_jobs.insert_one(doc)
    return doc


async def update_job_status(
    db, job_id: str, status: str, **extra
) -> None:
    update: Dict[str, Any] = {"status": status, "updated_at": _now()}
    if status in ("completed", "failed"):
        update["completed_at"] = _now()
    update.update(extra)
    await db.import_jobs.update_one({"_id": job_id}, {"$set": update})


async def save_import_errors(db, job_id: str, errors: List[Dict[str, Any]]) -> None:
    if not errors:
        return
    docs = [{"_id": str(uuid.uuid4()), "job_id": job_id, **e} for e in errors]
    await db.import_errors.insert_many(docs)
